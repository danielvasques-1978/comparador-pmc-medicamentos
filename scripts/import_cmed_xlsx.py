from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


PMC_COLUMNS = {
    "17": "PMC 17 %",
    "18": "PMC 18 %",
    "19": "PMC 19 %",
    "19.5": "PMC 19,5 %",
    "20": "PMC 20 %",
    "20.5": "PMC 20,5 %",
    "22.5": "PMC 22,5 %",
    "23": "PMC 23 %",
}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_price(value: object) -> float | None:
    if value is None or clean(value) in {"", "-"}:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    # A CMED marks ICMS-exempt presentations with an asterisk after the price.
    normalized = clean(value).rstrip("*").replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return round(float(normalized), 2)
    except ValueError:
        return None


def find_header_row(sheet) -> int:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if clean(row[0]).upper() == "SUBSTÂNCIA":
            return row_number
    raise ValueError("Cabeçalho SUBSTÂNCIA não encontrado na planilha CMED.")


def extract_table_date(sheet) -> str:
    for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
        for value in row:
            match = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", clean(value))
            if match:
                return match.group(1)
    raise ValueError("Data de publicação não encontrada na planilha CMED.")


def import_cmed(input_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = find_header_row(sheet)
    table_date = extract_table_date(sheet)
    headers = [clean(value) for value in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    columns = {header: index for index, header in enumerate(headers)}

    required = [
        "SUBSTÂNCIA",
        "LABORATÓRIO",
        "CÓDIGO GGREM",
        "REGISTRO",
        "PRODUTO",
        "APRESENTAÇÃO",
        "TIPO DE PRODUTO (STATUS DO PRODUTO)",
        "COMERCIALIZAÇÃO 2025",
        *PMC_COLUMNS.values(),
    ]
    missing = [column for column in required if column not in columns]
    if missing:
        raise ValueError(f"Colunas obrigatórias ausentes: {', '.join(missing)}")

    medicines: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        prices = {rate: parse_price(row[columns[column]]) for rate, column in PMC_COLUMNS.items()}
        if not any(price is not None for price in prices.values()):
            continue

        ggrem_code = clean(row[columns["CÓDIGO GGREM"]])
        product_type = clean(row[columns["TIPO DE PRODUTO (STATUS DO PRODUTO)"]]) or "Não informado"
        medicines.append(
            {
                "id": ggrem_code,
                "name": clean(row[columns["PRODUTO"]]),
                "activeIngredient": clean(row[columns["SUBSTÂNCIA"]]),
                "laboratory": clean(row[columns["LABORATÓRIO"]]),
                "kind": product_type,
                "productType": product_type,
                "presentation": clean(row[columns["APRESENTAÇÃO"]]),
                "pmc": prices,
                "ggremCode": ggrem_code,
                "registration": clean(row[columns["REGISTRO"]]),
                "commercialized": clean(row[columns["COMERCIALIZAÇÃO 2025"]]).casefold() == "sim",
                "sourcePage": 0,
                "source": "CMED/Anvisa",
                "tableDate": table_date,
            }
        )

    ids = [medicine["id"] for medicine in medicines]
    if len(ids) != len(set(ids)):
        raise ValueError("A planilha contém códigos GGREM duplicados entre as apresentações com PMC.")
    return medicines


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python scripts/import_cmed_xlsx.py <entrada.xlsx> <saida.json>")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    medicines = import_cmed(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(medicines, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    commercialized = sum(bool(item["commercialized"]) for item in medicines)
    print(f"Importadas {len(medicines)} apresentações com PMC; {commercialized} comercializadas em 2025.")


if __name__ == "__main__":
    main()
